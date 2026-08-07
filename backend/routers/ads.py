from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from typing import List
from database import get_db
import io

import openpyxl
from fastapi import File, UploadFile
from fastapi.responses import StreamingResponse
from models import AdData, Video
from schemas import AdDataCreate, AdDataOut

router = APIRouter()


@router.get("/", response_model=List[AdDataOut])
def list_ad_data(skip: int = 0, limit: int = 500, db: Session = Depends(get_db)):
    return db.query(AdData).order_by(AdData.id.asc()).offset(skip).limit(limit).all()


@router.get("/{ad_id:int}", response_model=AdDataOut)
def get_ad_data(ad_id: int, db: Session = Depends(get_db)):
    item = db.query(AdData).filter(AdData.id == ad_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="投流数据不存在")
    return item


@router.post("/", response_model=AdDataOut)
def create_ad_data(item: AdDataCreate, db: Session = Depends(get_db)):
    data = item.model_dump()
    if data.get("impressions") and data.get("clicks"):
        data["ctr"] = round(data["clicks"] / data["impressions"] * 100, 2)
    if data.get("spend") and data.get("revenue"):
        data["roi"] = round(data["revenue"] / data["spend"], 2) if data["spend"] > 0 else 0
    db_item = AdData(**data)
    db.add(db_item)
    db.commit()
    db.refresh(db_item)
    # 新增投流数据后异步触发复盘智能体（不阻塞本次请求）
    from ai_workflow.agent_graph import trigger_ad_review_background
    trigger_ad_review_background()
    return db_item


@router.put("/{ad_id:int}", response_model=AdDataOut)
def update_ad_data(ad_id: int, item: AdDataCreate, db: Session = Depends(get_db)):
    db_item = db.query(AdData).filter(AdData.id == ad_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="投流数据不存在")
    for key, value in item.model_dump(exclude_unset=True).items():
        setattr(db_item, key, value)
    db.commit()
    db.refresh(db_item)
    return db_item


@router.delete("/{ad_id:int}")
def delete_ad_data(ad_id: int, db: Session = Depends(get_db)):
    db_item = db.query(AdData).filter(AdData.id == ad_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="投流数据不存在")
    db.delete(db_item)
    db.commit()
    return {"message": "删除成功"}


from pydantic import BaseModel


class BatchDeleteIn(BaseModel):
    ids: list[int] = []


@router.post("/batch_delete")
def batch_delete_ads(data: BatchDeleteIn, db: Session = Depends(get_db)):
    ids = data.ids
    if not ids:
        raise HTTPException(status_code=400, detail="请先勾选要删除的投流数据")
    items = db.query(AdData).filter(AdData.id.in_(ids)).all()
    if not items:
        raise HTTPException(status_code=400, detail="没有可删除的投流数据")
    db.query(AdData).filter(AdData.id.in_(ids)).delete(synchronize_session=False)
    db.commit()
    return {"message": f"已删除 {len(items)} 条投流数据记录"}


@router.get("/view/high_priority")
def high_priority_view(db: Session = Depends(get_db)):
    """高优先级问题视图：ROI < 1 或有异常的数据"""
    items = db.query(AdData).filter(
        (AdData.roi < 1) | (AdData.anomaly.isnot(None))
    ).all()
    return items


# Excel 表头别名 → 投流字段（与千川素材报表导出口径对齐）
IMPORT_HEADER_MAP = {
    "投放日期": "ad_date",
    "日期": "ad_date",
    "视频编号": "video_code",
    "素材编号": "video_code",
    "视频": "video_code",
    "计划名称": "plan_name",
    "计划": "plan_name",
    "内容方向": "content_direction",
    "消耗": "spend",
    "花费": "spend",
    "展现": "impressions",
    "展现量": "impressions",
    "曝光": "impressions",
    "点击": "clicks",
    "点击量": "clicks",
    "购物车点击": "cart_clicks",
    "加购": "cart_clicks",
    "成交金额": "revenue",
    "成交额": "revenue",
    "GMV": "revenue",
    "订单数": "orders",
    "订单量": "orders",
    "成交订单": "orders",
    "播放量": "play_count",
    "播放": "play_count",
    "2秒跳出率": "bounce_rate_2s",
    "2s跳出率": "bounce_rate_2s",
    "5秒完播率": "completion_rate_5s",
    "5s完播率": "completion_rate_5s",
    "完播率": "completion_rate",
    "负责人": "owner",
    "状态": "status",
    "反馈": "feedback",
    "用户反馈": "feedback",
}

IMPORT_TEMPLATE_HEADERS = [
    "投放日期", "视频编号", "计划名称", "内容方向", "消耗", "展现", "点击",
    "购物车点击", "成交金额", "订单数", "播放量", "2秒跳出率", "5秒完播率",
    "完播率", "负责人", "状态", "反馈",
]


def _to_float(value) -> float | None:
    if value is None:
        return None
    text = str(value).replace(",", "").replace("，", "").strip().rstrip("%")
    if not text:
        return None
    try:
        return round(float(text), 2)
    except ValueError:
        return None


def _to_int(value) -> int | None:
    f = _to_float(value)
    return int(f) if f is not None else None


@router.get("/import_template")
def import_template():
    """下载投流数据批量导入 Excel 模板（字段与千川素材报表口径一致）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "投流导入"
    ws.append(IMPORT_TEMPLATE_HEADERS)
    ws.append(["2026-08-06", "V001", "测试计划-早餐果汁", "上班族早餐", 300, 12000, 960,
               90, 712, 8, 12000, 38, 22, 15, "张三", "投放中", "评论集中问清洗麻烦吗"])
    ws.append(["2026-08-06", "V002", "测试计划-健身饮品", "健身饮品", 200, 8000, 480,
               35, 178, 2, 8000, 51, 16, 9, "张三", "投放中", "前3秒钩子偏弱"])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ad_import_template.xlsx"},
    )


@router.post("/import")
async def import_ad_data(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Excel 批量导入投流数据：逐行写入投流表，CTR/ROI 自动计算，视频编号自动关联视频任务"""
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 格式的 Excel 文件")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(await file.read()), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Excel 解析失败，请检查文件格式")
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        raise HTTPException(status_code=400, detail="文件为空或缺少表头")

    col_map = {}
    for idx, h in enumerate(header):
        key = IMPORT_HEADER_MAP.get(str(h).strip()) if h is not None else None
        if key:
            col_map[idx] = key
    if not col_map:
        raise HTTPException(status_code=400, detail="表头无法识别，请使用下载的导入模板")

    # 视频编号 → video_id 映射缓存
    video_cache = {}
    for code, in db.query(Video.video_code).all():
        if code:
            video_cache[str(code).strip().upper()] = code

    results = []
    imported = 0
    skipped = 0
    for row in rows_iter:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        item = {}
        for idx, field in col_map.items():
            if idx < len(row):
                item[field] = row[idx]
        plan_name = str(item.get("plan_name") or "").strip()
        if not plan_name:
            skipped += 1
            results.append({"plan_name": "(无计划名称)", "status": "跳过", "message": "计划名称为空"})
            continue

        video_code = str(item.get("video_code") or "").strip().upper()
        video_id = None
        if video_code and video_code in video_cache:
            video = db.query(Video).filter(Video.video_code == video_code).first()
            video_id = video.id if video else None

        spend = _to_float(item.get("spend")) or 0
        impressions = _to_int(item.get("impressions")) or 0
        clicks = _to_int(item.get("clicks")) or 0
        revenue = _to_float(item.get("revenue")) or 0

        db_item = AdData(
            video_id=video_id,
            ad_date=str(item.get("ad_date") or "").strip() or None,
            plan_name=plan_name,
            content_direction=str(item.get("content_direction") or "").strip() or None,
            spend=spend,
            impressions=impressions,
            clicks=clicks,
            cart_clicks=_to_int(item.get("cart_clicks")) or 0,
            revenue=revenue,
            orders=_to_int(item.get("orders")) or 0,
            play_count=_to_int(item.get("play_count")) or 0,
            bounce_rate_2s=_to_float(item.get("bounce_rate_2s")) or 0,
            completion_rate_5s=_to_float(item.get("completion_rate_5s")) or 0,
            completion_rate=_to_float(item.get("completion_rate")) or 0,
            ctr=round(clicks / impressions * 100, 2) if impressions > 0 else 0,
            roi=round(revenue / spend, 2) if spend > 0 else 0,
            owner=str(item.get("owner") or "").strip() or None,
            status=str(item.get("status") or "").strip() or "投放中",
            feedback=str(item.get("feedback") or "").strip() or None,
            priority="P1",
        )
        db.add(db_item)
        db.flush()
        imported += 1
        results.append({
            "plan_name": plan_name,
            "ad_date": db_item.ad_date or "-",
            "video_code": video_code or "-",
            "spend": db_item.spend,
            "roi": db_item.roi,
            "status": "导入成功",
            "message": "" if video_id else "未匹配到视频任务，已留空关联",
        })
    db.commit()
    # 导入完成后异步触发复盘智能体，自动为这批新数据生成复盘结论
    from ai_workflow.agent_graph import trigger_ad_review_background
    if imported > 0:
        trigger_ad_review_background()
    return {
        "message": f"导入完成：成功 {imported} 条，跳过 {skipped} 条",
        "imported": imported,
        "skipped": skipped,
        "results": results,
    }

