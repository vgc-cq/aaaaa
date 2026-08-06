import csv
import io
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from fastapi import APIRouter, Body, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
from database import get_db
from models import Content, Product, Script
from schemas import ProductCreate, ProductOut
from ai_workflow.workflow import call_ai, parse_ai_response
from services.scoring import score_product, status_for_score

router = APIRouter()

PRODUCT_FIELDS = [
    "product_code", "name", "category", "price_min", "price_max", "commission",
    "sales_heat", "reputation", "target_users", "selling_points", "pain_points",
    "risk_words", "score", "owner", "status",
]

# Excel 表头别名 → 商品字段（与商品库表单字段保持一致）
IMPORT_HEADER_MAP = {
    "编号": "product_code",
    "商品编号": "product_code",
    "商品名称": "name",
    "类目": "category",
    "负责人": "owner",
    "最低价": "price_min",
    "最高价": "price_max",
    "佣金": "commission",
    "热度": "sales_heat",
    "销量": "sales_heat",
    "销量/热度": "sales_heat",
    "口碑": "reputation",
    "好评率": "reputation",
    "目标人群": "target_users",
    "卖点": "selling_points",
    "痛点": "pain_points",
    "风险词": "risk_words",
    "评分": "score",
    "状态": "status",
    "价格区间": "price_range",
}

IMPORT_TEMPLATE_HEADERS = [
    "编号", "商品名称", "类目", "价格区间", "佣金", "热度", "口碑",
    "状态", "负责人", "卖点",
]


def _first_text(text: str, names: list[str]) -> str:
    for name in names:
        m = re.search(rf"{name}\s*[?:]\s*([^?;\n\r]+)", text, re.I)
        if m:
            return m.group(1).strip()
    return ""


def _num(value) -> float | None:
    m = re.search(r"\d+(?:\.\d+)?", str(value or ""))
    return float(m.group(0)) if m else None


def _parse_price(text: str) -> tuple[float | None, float | None]:
    m = re.search(r"(?:??|??|????)\s*[?:]?\s*??\s*(\d+(?:\.\d+)?)\s*[-~???]\s*??\s*(\d+(?:\.\d+)?)", text, re.I)
    if m:
        return float(m.group(1)), float(m.group(2))
    m = re.search(r"(?:??|??)\s*[?:]?\s*??\s*(\d+(?:\.\d+)?)", text, re.I)
    if m:
        v = float(m.group(1))
        return v, v
    return None, None


def _next_product_code(db: Session) -> str:
    rows = db.query(Product.product_code).filter(Product.product_code.like("P%")).all()
    max_num = 0
    for (code,) in rows:
        digits = "".join(ch for ch in str(code or "") if ch.isdigit())
        if digits:
            max_num = max(max_num, int(digits))
    return f"P{max_num + 1:03d}"


def _parse_price_range(text: str) -> tuple[float | None, float | None]:
    nums = re.findall(r"\d+(?:\.\d+)?", str(text or ""))
    if not nums:
        return None, None
    vals = [float(n) for n in nums]
    return min(vals), max(vals)


def _score_text_signal(text: str | None, good_words: list[str], max_score: float) -> float:
    if not text:
        return 0
    hit = sum(1 for word in good_words if word in text)
    return min(max_score, hit / max(1, len(good_words)) * max_score)


def calculate_product_score(data: dict) -> float:
    """???????0-100???? 25 + ?? 25 + ?? 20 + ??/?? 20 - ?? 10?"""
    score = 0.0
    commission = _num(data.get("commission")) or 0
    score += min(25, commission / 30 * 25)

    heat_text = str(data.get("sales_heat") or "")
    heat_num = _num(heat_text) or 0
    if any(k in heat_text for k in ["?", "?", "?", "?"]):
        heat_num *= 10 if "?" in heat_text else 1.5
    score += min(25, heat_num / 10000 * 25)

    rep_text = str(data.get("reputation") or "")
    rep_num = _num(rep_text) or 0
    if "??" in rep_text and rep_num > 10:
        score += min(20, rep_num / 100 * 20)
    elif rep_num:
        score += min(20, rep_num / 5 * 20)
    else:
        score += _score_text_signal(rep_text, ["??", "??", "??", "??"], 20)

    score += _score_text_signal(str(data.get("selling_points") or ""), ["??", "??", "??", "??", "??", "???", "??"], 12)
    score += min(8, len(str(data.get("target_users") or "")) / 30 * 8)

    risk = str(data.get("risk_words") or "")
    penalty = 0
    if risk.strip():
        penalty = min(10, 2 + len([x for x in re.split(r"[?,??;\s]+", risk) if x]) * 2)
    return round(max(0, min(100, score - penalty)), 1)


def _heuristic_parse_product(text: str, db: Session) -> dict:
    price_min, price_max = _parse_price(text)
    return {
        "product_code": _first_text(text, ["????", "??"]) or _next_product_code(db),
        "name": _first_text(text, ["????", "???", "????", "??"]),
        "category": _first_text(text, ["??", "??", "??"]),
        "price_min": price_min,
        "price_max": price_max,
        "commission": _num(_first_text(text, ["??", "????"])),
        "sales_heat": _first_text(text, ["??", "??", "??/??"]),
        "reputation": _first_text(text, ["??", "??", "???"]),
        "target_users": _first_text(text, ["????", "????", "??"]),
        "selling_points": _first_text(text, ["??", "????", "?????"]),
        "pain_points": _first_text(text, ["??", "????"]),
        "risk_words": _first_text(text, ["???", "????", "???"]),
        "score": _num(_first_text(text, ["??", "????"])),
        "owner": _first_text(text, ["???", "???"]) or "????",
        "status": _first_text(text, ["??"]) or "???",
    }


def _normalize_product(data: dict, text: str, db: Session) -> dict:
    fallback = _heuristic_parse_product(text, db)
    result = {}
    for key in PRODUCT_FIELDS:
        value = data.get(key, None) if isinstance(data, dict) else None
        result[key] = value if value not in (None, "") else fallback.get(key)
    for key in ["price_min", "price_max", "commission", "score"]:
        result[key] = _num(result.get(key))
    result["product_code"] = result.get("product_code") or _next_product_code(db)
    result["status"] = result.get("status") or "???"
    if result.get("score") in (None, 0):
        result["score"] = calculate_product_score(result)
    return result


def _ai_parse_product(text: str, db: Session) -> dict:
    prompt = f"""
????????????????????????????????? JSON??? Markdown?
???product_code,name,category,price_min,price_max,commission,sales_heat,reputation,target_users,selling_points,pain_points,risk_words,score,owner,status
??????????? null?
?????{text[:12000]}
"""
    parsed = parse_ai_response(call_ai(prompt))
    if parsed.get("local_demo") or parsed.get("error") or parsed.get("parse_error"):
        parsed = {}
    return _normalize_product(parsed, text, db)


def _read_txt(raw: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            return raw.decode(enc)
        except Exception:
            pass
    return raw.decode("utf-8", errors="ignore")


def _read_docx(raw: bytes) -> str:
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as zf:
            xml = zf.read("word/document.xml")
        root = ET.fromstring(xml)
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        paragraphs = []
        for p in root.findall(".//w:p", ns):
            texts = [t.text or "" for t in p.findall(".//w:t", ns)]
            if texts:
                paragraphs.append("".join(texts))
        return "\n".join(paragraphs)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"docx ?????{str(e)}")


def _read_xlsx(raw: bytes) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), data_only=True)
        lines = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                vals = [str(v) for v in row if v is not None]
                if vals:
                    lines.append("?".join(vals))
        return "\n".join(lines)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"xlsx ?????{str(e)}")


def _read_csv(raw: bytes) -> str:
    text = _read_txt(raw)
    rows = csv.reader(io.StringIO(text))
    return "\n".join("?".join(cell for cell in row if cell) for row in rows)


def _extract_document_text(filename: str, raw: bytes) -> str:
    ext = Path(filename).suffix.lower()
    if ext == ".txt":
        return _read_txt(raw)
    if ext == ".docx":
        return _read_docx(raw)
    if ext == ".xlsx":
        return _read_xlsx(raw)
    if ext == ".csv":
        return _read_csv(raw)
    raise HTTPException(status_code=400, detail="??? txt/docx/xlsx/csv ??")


def _ai_analyze_product(product: Product) -> dict | None:
    """调用 DeepSeek 选品分析（未配置 Key 或失败时返回 None，评分仍按规则执行）"""
    price = f"{product.price_min}-{product.price_max}元" if product.price_min is not None else ""
    prompt = f"""请分析以下商品的选品价值：

商品名称：{product.name or ''}
价格：{price}
销量/热度：{product.sales_heat or ''}
口碑评价：{product.reputation or ''}
类目：{product.category or ''}
佣金比例：{product.commission or ''}
目标人群：{product.target_users or ''}
已整理卖点：{product.selling_points or ''}
已整理痛点：{product.pain_points or ''}
风险词约束：{product.risk_words or ''}

请输出JSON格式的选品分析报告，包含：score(0-100评分)、target_users(适合人群)、risk_points(风险点)、content_angles(内容角度建议)、recommendation(是否建议测试及理由)。
要求：基于提供的销量/热度、口碑、佣金等数据做判断；数据不足时明确说明"数据不足"并降低相应维度评分，不得编造未提供的市场数据；风险词约束必须体现在风险点中。"""
    result = parse_ai_response(call_ai(prompt))
    if result.get("local_demo") or result.get("error") or result.get("parse_error"):
        return None
    return result


@router.get("/", response_model=List[ProductOut])
def list_products(skip: int = 0, limit: int = 500, status: str = None, db: Session = Depends(get_db)):
    query = db.query(Product)
    if status:
        query = query.filter(Product.status == status)
    return query.order_by(Product.id.asc()).offset(skip).limit(limit).all()


@router.get("/import_template")
def import_template():
    """下载商品批量导入 Excel 模板（字段与商品库表单一致）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "商品导入"
    ws.append(IMPORT_TEMPLATE_HEADERS)
    ws.append(["P002", "便携无线榨汁杯300ml", "厨房小家电", "79-129元", 15,
               "月销5000+", "4.8分/好评率96%", "待评估", "张三", "便携、无线、易清洗"])
    ws.append(["", "便携折叠硅胶水杯", "日用百货", "29-59元", 20,
               "月销8000+", "4.7分/好评率95%", "待评估", "李四", "可折叠、轻便、大容量"])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=product_import_template.xlsx"},
    )


@router.post("/import")
async def import_products(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Excel 批量导入商品：逐行读取，直接写入商品库（编号留空自动生成，同名跳过）"""
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 格式的 Excel 文件")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(await file.read()), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Excel 解析失败，请检查文件格式")
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        raise HTTPException(status_code=400, detail="文件为空或缺少表头")

    col_map = {}
    for idx, h in enumerate(header):
        key = re.sub(r"[\s%]", "", str(h or ""))
        if key in IMPORT_HEADER_MAP:
            col_map[idx] = IMPORT_HEADER_MAP[key]
    if not col_map:
        raise HTTPException(status_code=400, detail="表头无法识别，请使用下载的导入模板")

    results = []
    imported = 0
    skipped = 0
    for row in rows_iter:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        item = {}
        for idx, field in col_map.items():
            if idx < len(row) and row[idx] is not None:
                item[field] = str(row[idx]).strip()

        name = item.get("name", "")
        if not name:
            skipped += 1
            results.append({"imported": False, "name": name, "error": "缺少商品名称"})
            continue
        if db.query(Product).filter(Product.name == name).first():
            skipped += 1
            results.append({"imported": False, "name": name, "error": "同名商品已存在，跳过"})
            continue

        code = item.get("product_code") or _next_product_code(db)
        if db.query(Product).filter(Product.product_code == code).first():
            code = _next_product_code(db)

        price_min = _num(item.get("price_min"))
        price_max = _num(item.get("price_max"))
        if item.get("price_range"):
            pm, px = _parse_price_range(item["price_range"])
            if price_min is None:
                price_min = pm
            if price_max is None:
                price_max = px

        product = Product(
            product_code=code,
            name=name,
            category=item.get("category"),
            owner=item.get("owner") or "选品运营",
            price_min=price_min,
            price_max=price_max,
            commission=_num(item.get("commission")),
            sales_heat=item.get("sales_heat"),
            reputation=item.get("reputation"),
            target_users=item.get("target_users"),
            selling_points=item.get("selling_points"),
            pain_points=item.get("pain_points"),
            risk_words=item.get("risk_words"),
            score=_num(item.get("score")),
            status=item.get("status") or "待评估",
        )
        db.add(product)
        db.flush()
        imported += 1
        results.append({
            "imported": True,
            "product_code": code,
            "name": name,
            "score": product.score,
            "status": product.status,
            "error": None,
        })

    db.commit()
    return {"total": len(results), "imported": imported, "skipped": skipped, "results": results}


@router.post("/ai_select")
def ai_select(data: dict = Body(...), db: Session = Depends(get_db)):
    """AI 选品：勾选商品后按规则评分回填，状态按分数自动联动，DeepSeek 辅助分析"""
    ids = data.get("product_ids") or []
    query = db.query(Product)
    if ids:
        query = query.filter(Product.id.in_(ids))
    products = query.all()
    if not products:
        raise HTTPException(status_code=400, detail="请先勾选要分析的商品")

    results = []
    for p in products:
        rule = score_product(p)
        ai = _ai_analyze_product(p)
        p.score = rule["total"]
        p.status = status_for_score(rule["total"])
        if ai:
            # AI 只补全空字段，补完锁定：保证评分字段稳定，重复 AI 选品分数不再变化
            if not str(p.target_users or "").strip():
                t = ai.get("target_users")
                if isinstance(t, list):
                    p.target_users = "、".join(str(x) for x in t)
                elif isinstance(t, str) and t.strip():
                    p.target_users = t
            if not str(p.risk_words or "").strip():
                rp = ai.get("risk_points")
                if isinstance(rp, list):
                    p.risk_words = "、".join(str(x) for x in rp)
                elif isinstance(rp, str) and rp.strip():
                    p.risk_words = rp
        db.flush()
        results.append({
            "id": p.id,
            "product_code": p.product_code,
            "name": p.name,
            "score": p.score,
            "status": p.status,
            "dimensions": rule["dimensions"],
            "ai": ai,
        })

    db.commit()
    return {"total": len(products), "results": results}


@router.post("/import_document")
async def import_product_document(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """?? txt/docx/xlsx/csv??????????????????????"""
    filename = file.filename or ""
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="??????")
    text = _extract_document_text(filename, raw)
    if not text.strip():
        raise HTTPException(status_code=400, detail="???????????")
    return {"filename": filename, "text_preview": text[:1000], "parsed": _ai_parse_product(text, db)}


@router.post("/score")
def preview_product_score(product: ProductCreate):
    return {"score": calculate_product_score(product.model_dump())}


@router.get("/{product_id:int}", response_model=ProductOut)
def get_product(product_id: int, db: Session = Depends(get_db)):
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="?????")
    return product


@router.post("/", response_model=ProductOut)
def create_product(product: ProductCreate, db: Session = Depends(get_db)):
    data = product.model_dump()
    if data.get("score") in (None, 0):
        data["score"] = calculate_product_score(data)
    db_product = Product(**data)
    db.add(db_product)
    db.commit()
    db.refresh(db_product)
    return db_product


@router.put("/{product_id:int}", response_model=ProductOut)
def update_product(product_id: int, product: ProductCreate, db: Session = Depends(get_db)):
    db_product = db.query(Product).filter(Product.id == product_id).first()
    if not db_product:
        raise HTTPException(status_code=404, detail="?????")
    data = product.model_dump(exclude_unset=True)
    for key, value in data.items():
        setattr(db_product, key, value)
    db.commit()
    db.refresh(db_product)
    return db_product


@router.delete("/{product_id:int}")
def delete_product(product_id: int, db: Session = Depends(get_db)):
    db_product = db.query(Product).filter(Product.id == product_id).first()
    if not db_product:
        raise HTTPException(status_code=404, detail="?????")
    db.delete(db_product)
    db.commit()
    return {"message": "????"}


@router.post("/batch_delete")
def batch_delete_products(data: dict = Body(...), db: Session = Depends(get_db)):
    """批量删除商品及其关联的内容拆解、脚本，避免孤儿数据"""
    ids = data.get("product_ids") or []
    if not ids:
        raise HTTPException(status_code=400, detail="请先勾选要删除的商品")
    products = db.query(Product).filter(Product.id.in_(ids)).all()
    if not products:
        raise HTTPException(status_code=400, detail="没有可删除的商品")

    db.query(Content).filter(Content.product_id.in_(ids)).delete(synchronize_session=False)
    db.query(Script).filter(Script.product_id.in_(ids)).delete(synchronize_session=False)
    db.query(Product).filter(Product.id.in_(ids)).delete(synchronize_session=False)
    db.commit()
    return {"message": f"已删除 {len(products)} 个商品及其关联内容/脚本"}


@router.get("/view/today")
def today_tasks(db: Session = Depends(get_db)):
    """???????"""
    return db.query(Product).filter(Product.status.in_(["???", "???"])).all()


@router.get("/view/kanban")
def kanban_view(db: Session = Depends(get_db)):
    """???????"""
    products = db.query(Product).all()
    kanban = {}
    for p in products:
        kanban.setdefault(p.status, []).append({"id": p.id, "name": p.name, "score": p.score, "owner": p.owner})
    return kanban
